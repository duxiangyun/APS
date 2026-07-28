import xlrd
import openpyxl

def read_xls(file_path):
    book = xlrd.open_workbook(file_path)
    sheets = {}
    for sheet in book.sheets():
        data = []
        for row_idx in range(sheet.nrows):
            row = []
            for col_idx in range(sheet.ncols):
                val = sheet.cell_value(row_idx, col_idx)
                if isinstance(val, float) and val == int(val):
                    val = int(val)
                row.append(val)
            data.append(row)
        sheets[sheet.name.strip()] = data
    return sheets

def read_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            row_data = []
            for val in row:
                if isinstance(val, float) and val == int(val):
                    val = int(val)
                row_data.append(val)
            data.append(row_data)
        sheets[sheet_name.strip()] = data
    return sheets

def compare_sheet(name, data1, data2):
    rows1 = len(data1)
    rows2 = len(data2)
    
    if rows1 != rows2:
        print(f'❌ {name}: 行数不一致 - 原始Excel: {rows1}, 导出Excel: {rows2}')
        return False
    
    all_match = True
    for i in range(rows1):
        row1 = data1[i]
        row2 = data2[i]
        
        cols1 = len(row1)
        cols2 = len(row2)
        
        if cols1 != cols2:
            print(f'❌ {name}[{i+1}]: 列数不一致 - 原始Excel: {cols1}, 导出Excel: {cols2}')
            all_match = False
            continue
        
        for j in range(cols1):
            val1 = row1[j]
            val2 = row2[j]
            
            if val1 == '' and val2 is None:
                continue
            
            if val1 is None and val2 == '':
                continue
            
            try:
                num1 = float(val1)
                num2 = float(val2)
                if abs(num1 - num2) > 0.0001:
                    print(f'❌ {name}[{i+1}][{j+1}]: 值不一致 - 原始: {val1}, 导出: {val2}')
                    all_match = False
                continue
            except (ValueError, TypeError):
                pass
            
            if val1 != val2:
                print(f'❌ {name}[{i+1}][{j+1}]: 值不一致 - 原始: {val1}, 导出: {val2}')
                all_match = False
    
    if all_match:
        print(f'✅ {name}: 数据完全一致 ({rows1}行 x {len(data1[0])}列)')
    return all_match

def main():
    print('=== Excel文件对比 ===')
    print()
    
    xls_path = '/Users/duxiangyun/PythonProjects/APS/Input/APS-JD.xls'
    xlsx_path = '/Users/duxiangyun/PythonProjects/APS/Excel2DB/aps_model_export.xlsx'
    
    print(f'原始Excel: {xls_path}')
    print(f'导出Excel: {xlsx_path}')
    print()
    
    xls_sheets = read_xls(xls_path)
    xlsx_sheets = read_xlsx(xlsx_path)
    
    print('--- Sheet名称对比 ---')
    xls_names = set(xls_sheets.keys())
    xlsx_names = set(xlsx_sheets.keys())
    
    only_in_xls = xls_names - xlsx_names
    only_in_xlsx = xlsx_names - xls_names
    
    if only_in_xls:
        print(f'❌ 仅在原始Excel中存在的Sheet: {only_in_xls}')
    if only_in_xlsx:
        print(f'❌ 仅在导出Excel中存在的Sheet: {only_in_xlsx}')
    
    common_sheets = xls_names & xlsx_names
    print(f'✅ 共同的Sheet: {len(common_sheets)}个')
    
    print()
    print('--- Sheet顺序对比 ---')
    print(f'原始Excel顺序: {list(xls_sheets.keys())}')
    print(f'导出Excel顺序: {list(xlsx_sheets.keys())}')
    
    if list(xls_sheets.keys()) == list(xlsx_sheets.keys()):
        print('✅ Sheet顺序一致')
    else:
        print('❌ Sheet顺序不一致')
    
    print()
    print('--- 数据内容对比 ---')
    all_passed = True
    
    for sheet_name in xls_sheets:
        if sheet_name in xlsx_sheets:
            xls_data = xls_sheets[sheet_name]
            xlsx_data = xlsx_sheets[sheet_name]
            
            print()
            all_passed = compare_sheet(sheet_name, xls_data, xlsx_data) and all_passed
    
    print()
    print('=======================')
    if all_passed:
        print('🎉 所有数据对比通过！两个Excel文件内容完全一致')
    else:
        print('❌ 部分数据对比失败，请检查上述错误信息')

if __name__ == '__main__':
    main()