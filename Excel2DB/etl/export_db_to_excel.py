import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def export_to_excel(db_path, output_path):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    
    wb = openpyxl.Workbook()
    
    ws = wb.active
    ws.title = '综合表'
    
    ws.cell(row=1, column=1, value='')
    ws.cell(row=1, column=2, value='**本数据加入在制品数据需求')
    for col in range(3, 7):
        ws.cell(row=1, column=col, value='')
    
    ws.cell(row=2, column=1, value='')
    ws.cell(row=2, column=5, value='读入数据表')
    
    config_layout = [
        (3, 'nfixtable', '工装表'),
        (4, 'nouttable', '外协表'),
        (5, 'nrawlimtable', '采购限制表'),
        (6, 'nsubtable', '替代关系表'),
        (7, 'nwiptable', '在制品表'),
        (10, 'nperiod', '计划期长度'),
        (11, 'nbom', 'BOM表记录数'),
        (12, 'nrouting', '工艺路线表记录数'),
        (13, 'norder', '订单表记录数'),
        (14, 'nplant', '工厂数'),
        (15, 'nequip', '设备数'),
        (16, 'nfixture', '工装数'),
        (17, 'nproduct', '产品数'),
        (18, 'nselfmade', '自制产品数'),
        (19, 'nrawmat', '原料数'),
        (20, 'nordclass', '订单等级'),
        (21, 'dutytime', '每班时长（分钟）'),
        (22, 'dayshift', '每期班次数'),
        (23, 'npmaxt', '最多替代工艺数'),
        (24, 'maxpro', '最大加工周期'),
        (25, 'ndemrate', '需求率'),
        (26, 'nordelay', '订单最大允许延期'),
    ]
    
    for row, key, desc in config_layout:
        cursor = conn.execute('SELECT value FROM system_config WHERE key = ?', (key,))
        result = cursor.fetchone()
        value = result['value'] if result else ''
        
        ws.cell(row=row, column=5, value=desc)
        ws.cell(row=row, column=6, value=value)
    
    ws.cell(row=8, column=1, value='')
    for col in range(2, 7):
        ws.cell(row=8, column=col, value='')
    
    ws.cell(row=9, column=1, value='')
    for col in range(2, 7):
        ws.cell(row=9, column=col, value='')
    ws.cell(row=9, column=5, value='主要参数')
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 5
    
    def add_header(ws, row, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    ws = wb.create_sheet(title='订单表')
    ws.cell(row=1, column=1, value='')
    add_header(ws, 2, ['', '订单序号', '产品代码', '订单价格', '订单数量', '订单交期', '订单等级', '允许延期', '延期罚金', '生产提前期'])
    
    cursor = conn.execute('SELECT * FROM v_orders')
    rows = cursor.fetchall()
    for row_idx, row_data in enumerate(rows, 3):
        ws.cell(row=row_idx, column=1, value='')
        ws.cell(row=row_idx, column=2, value=row_data['no'])
        ws.cell(row=row_idx, column=3, value=row_data['prod_code'])
        ws.cell(row=row_idx, column=4, value=row_data['price'])
        ws.cell(row=row_idx, column=5, value=row_data['quantity'])
        ws.cell(row=row_idx, column=6, value=row_data['time'])
        ws.cell(row=row_idx, column=7, value=row_data['cls'])
        ws.cell(row=row_idx, column=8, value=row_data['delay'])
        ws.cell(row=row_idx, column=9, value=row_data['fine'])
        ws.cell(row=row_idx, column=10, value=0)
    
    for i in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    
    ws = wb.create_sheet(title='BOM')
    ws.cell(row=1, column=1, value='')
    add_header(ws, 2, ['', '序号', 'BOM层级', '父物料', '父物料名称', '子物料', '子物料名称', '数量'])
    
    cursor = conn.execute('SELECT * FROM v_bom')
    rows = cursor.fetchall()
    for row_idx, row_data in enumerate(rows, 3):
        ws.cell(row=row_idx, column=1, value=row_data['group_id'] if row_data['group_id'] else '')
        ws.cell(row=row_idx, column=2, value=row_data['seq'])
        ws.cell(row=row_idx, column=3, value=row_data['level'])
        ws.cell(row=row_idx, column=4, value=row_data['parent_code'])
        ws.cell(row=row_idx, column=5, value=row_data['parent_name'])
        ws.cell(row=row_idx, column=6, value=row_data['child_code'])
        ws.cell(row=row_idx, column=7, value=row_data['child_name'])
        ws.cell(row=row_idx, column=8, value=row_data['quantity'])
    
    for i in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15
    
    ws = wb.create_sheet(title='工艺路线')
    ws.cell(row=1, column=1, value='')
    add_header(ws, 2, ['', '序号', '物料代码', '物料名称', '多工艺', '设备代码', '生产线', '工序', '工装代码', '工装数量', 'MaxT', '1', '2', '3'])
    
    cursor = conn.execute('SELECT * FROM v_routing_excel')
    rows = cursor.fetchall()
    for row_idx, row_data in enumerate(rows, 3):
        ws.cell(row=row_idx, column=1, value=row_data['group_id'] if row_data['group_id'] else '')
        ws.cell(row=row_idx, column=2, value=row_data['seq'])
        ws.cell(row=row_idx, column=3, value=row_data['material_code'])
        ws.cell(row=row_idx, column=4, value=row_data['material_name'])
        ws.cell(row=row_idx, column=5, value=row_data['process_alt'])
        ws.cell(row=row_idx, column=6, value=row_data['equipment_code'])
        ws.cell(row=row_idx, column=7, value=row_data['production_line'])
        ws.cell(row=row_idx, column=8, value=row_data['stage_name'])
        ws.cell(row=row_idx, column=9, value=row_data['tooling_code'])
        ws.cell(row=row_idx, column=10, value=row_data['tooling_quantity'])
        ws.cell(row=row_idx, column=11, value=row_data['max_t'])
        ws.cell(row=row_idx, column=12, value=row_data['stage_1'])
        ws.cell(row=row_idx, column=13, value=row_data['stage_2'])
        ws.cell(row=row_idx, column=14, value=row_data['stage_3'])
    
    for i in range(1, 15):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    
    ws = wb.create_sheet(title='设备表')
    ws.cell(row=1, column=1, value='')
    add_header(ws, 2, ['', '序号', '设备代码', '单位成本', '设备数', '设备利用率', '加班率', '加班成本', '', '', '', '', '', '', '', '', '', '', ''])
    
    cursor = conn.execute('SELECT * FROM v_equipment')
    rows = cursor.fetchall()
    for row_idx, row_data in enumerate(rows, 3):
        ws.cell(row=row_idx, column=1, value='')
        ws.cell(row=row_idx, column=2, value=row_idx - 2)
        ws.cell(row=row_idx, column=3, value=row_data['code'])
        ws.cell(row=row_idx, column=4, value=row_data['cost'])
        ws.cell(row=row_idx, column=5, value=row_data['number'])
        ws.cell(row=row_idx, column=6, value=row_data['rate'])
        ws.cell(row=row_idx, column=7, value=row_data['overtime_rate'])
        ws.cell(row=row_idx, column=8, value=row_data['overtime'])
        for col in range(9, 18):
            ws.cell(row=row_idx, column=col, value='')
        ws.cell(row=row_idx, column=18, value=row_data['cost_t'])
        ws.cell(row=row_idx, column=19, value=row_data['cost_u'])
    
    for i in range(1, 20):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    
    ws = wb.create_sheet(title='工装表')
    ws.cell(row=1, column=1, value='')
    add_header(ws, 2, ['', '序号', '工装代码', '工装成本', '工装数', '工装利用率', '加班率', '加班成本'])
    
    cursor = conn.execute('SELECT * FROM v_tooling')
    rows = cursor.fetchall()
    for row_idx, row_data in enumerate(rows, 3):
        ws.cell(row=row_idx, column=1, value='')
        ws.cell(row=row_idx, column=2, value=row_idx - 2)
        ws.cell(row=row_idx, column=3, value=row_data['code'])
        ws.cell(row=row_idx, column=4, value=row_data['cost'])
        ws.cell(row=row_idx, column=5, value=row_data['quantity'])
        ws.cell(row=row_idx, column=6, value=row_data['rate'])
        ws.cell(row=row_idx, column=7, value=row_data['overtime'])
        ws.cell(row=row_idx, column=8, value=row_data['overtime_cost'])
    
    for i in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    
    ws = wb.create_sheet(title='产品表')
    ws.cell(row=1, column=1, value='')
    add_header(ws, 2, ['', '序号', '产品代码', '', '产品成本', '产品价格', '提前期', '初始库存', '期末库存', '最小库存', '最大库存', '库存成本'])
    
    cursor = conn.execute('SELECT * FROM v_product')
    rows = cursor.fetchall()
    for row_idx, row_data in enumerate(rows, 3):
        ws.cell(row=row_idx, column=1, value='')
        ws.cell(row=row_idx, column=2, value=row_idx - 2)
        ws.cell(row=row_idx, column=3, value=row_data['code'])
        ws.cell(row=row_idx, column=4, value=row_data['name'])
        ws.cell(row=row_idx, column=5, value=row_data['cost'])
        ws.cell(row=row_idx, column=6, value=row_data['price'])
        ws.cell(row=row_idx, column=7, value=row_data['lead_time'])
        ws.cell(row=row_idx, column=8, value=row_data['inv0'])
        ws.cell(row=row_idx, column=9, value=row_data['inv_t'])
        ws.cell(row=row_idx, column=10, value=row_data['inv_l'])
        ws.cell(row=row_idx, column=11, value=row_data['inv_u'])
        ws.cell(row=row_idx, column=12, value=row_data['inv_cost'])
    
    for i in range(1, 13):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    
    ws = wb.create_sheet(title='自制件')
    ws.cell(row=1, column=1, value='')
    add_header(ws, 2, ['', '序号', '自制件代码', '自制件名称', '虚拟件属性', '提前期', '初始库存', '期末库存', '最小库存', '最大库存', '库存成本'])
    
    cursor = conn.execute('SELECT * FROM v_semi')
    rows = cursor.fetchall()
    for row_idx, row_data in enumerate(rows, 3):
        ws.cell(row=row_idx, column=1, value='')
        ws.cell(row=row_idx, column=2, value=row_idx - 2)
        ws.cell(row=row_idx, column=3, value=row_data['code'])
        ws.cell(row=row_idx, column=4, value=row_data['name'])
        ws.cell(row=row_idx, column=5, value=row_data['dummy'])
        ws.cell(row=row_idx, column=6, value=row_data['lead_time'])
        ws.cell(row=row_idx, column=7, value=row_data['inv0'])
        ws.cell(row=row_idx, column=8, value=row_data['inv_t'])
        ws.cell(row=row_idx, column=9, value=row_data['inv_l'])
        ws.cell(row=row_idx, column=10, value=row_data['inv_u'])
        ws.cell(row=row_idx, column=11, value=row_data['inv_cost'])
    
    for i in range(1, 12):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    
    ws = wb.create_sheet(title='原材料表')
    ws.cell(row=1, column=1, value='')
    add_header(ws, 2, ['序号', '原料代码', '', '采购成本', '采购提前期', '初始库存', '期末库存', '最小库存', '最大库存', '库存成本'])
    
    cursor = conn.execute('SELECT * FROM v_raw')
    rows = cursor.fetchall()
    for row_idx, row_data in enumerate(rows, 3):
        ws.cell(row=row_idx, column=1, value=row_idx - 2)
        ws.cell(row=row_idx, column=2, value=row_data['code'])
        ws.cell(row=row_idx, column=3, value=row_data['name'])
        ws.cell(row=row_idx, column=4, value=row_data['cost'])
        ws.cell(row=row_idx, column=5, value=row_data['lead_time'])
        ws.cell(row=row_idx, column=6, value=row_data['inv0'])
        ws.cell(row=row_idx, column=7, value=row_data['inv_t'])
        ws.cell(row=row_idx, column=8, value=row_data['inv_l'])
        ws.cell(row=row_idx, column=9, value=row_data['inv_u'])
        ws.cell(row=row_idx, column=10, value=row_data['inv_cost'])
    
    for i in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    
    ws = wb.create_sheet(title='在制品')
    ws.cell(row=1, column=1, value='')
    for col in range(2, 7):
        ws.cell(row=1, column=col, value='')
    ws.cell(row=1, column=7, value='与ProMaxT相同')
    for col in range(8, 10):
        ws.cell(row=1, column=col, value='')
    add_header(ws, 2, ['', '序号', '在制品种类代码', '自制品种类', '在制件代码', '在制件名称', '总制造期', '已完成阶段', '在制数量'])
    
    cursor = conn.execute('SELECT * FROM v_wip')
    rows = cursor.fetchall()
    for row_idx, row_data in enumerate(rows, 3):
        ws.cell(row=row_idx, column=1, value='')
        ws.cell(row=row_idx, column=2, value=row_idx - 2)
        ws.cell(row=row_idx, column=3, value=2)
        ws.cell(row=row_idx, column=4, value='自制件')
        ws.cell(row=row_idx, column=5, value=row_data['material_code'])
        ws.cell(row=row_idx, column=6, value=row_data['material_name'])
        ws.cell(row=row_idx, column=7, value=row_data['max_t'])
        ws.cell(row=row_idx, column=8, value=row_data['stage'])
        ws.cell(row=row_idx, column=9, value=row_data['quantity'])
    
    for i in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    
    ws = wb.create_sheet(title='外协')
    ws.cell(row=1, column=1, value='')
    add_header(ws, 2, ['', '序号', '物料代码', '外协价格'] + [str(i) for i in range(1, 29)] + ['合计'])
    
    cursor = conn.execute('SELECT * FROM v_outsource')
    rows = cursor.fetchall()
    for row_idx, row_data in enumerate(rows, 3):
        ws.cell(row=row_idx, column=1, value='')
        ws.cell(row=row_idx, column=2, value=row_idx - 2)
        ws.cell(row=row_idx, column=3, value=row_data['code'])
        ws.cell(row=row_idx, column=4, value=row_data['cost'])
        total = 0
        for i in range(1, 29):
            col_name = f'quantity_t{i}'
            val = row_data[col_name] if col_name in row_data.keys() else 0
            ws.cell(row=row_idx, column=4 + i, value=val)
            total += val
        ws.cell(row=row_idx, column=33, value=total)
    
    for i in range(1, 34):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 8
    
    ws = wb.create_sheet(title='采购限制')
    ws.cell(row=1, column=1, value='')
    add_header(ws, 2, ['', '序号', '原料代码', ''] + [str(i) for i in range(1, 29)] + ['合计'])
    
    cursor = conn.execute('SELECT * FROM v_purchase_limit')
    rows = cursor.fetchall()
    for row_idx, row_data in enumerate(rows, 3):
        ws.cell(row=row_idx, column=1, value='')
        ws.cell(row=row_idx, column=2, value=row_idx - 2)
        ws.cell(row=row_idx, column=3, value=row_data['material_code'])
        ws.cell(row=row_idx, column=4, value=row_data['material_name'])
        total = 0
        for i in range(1, 29):
            col_name = f'quantity_t{i}'
            val = row_data[col_name] if col_name in row_data.keys() else 0
            ws.cell(row=row_idx, column=4 + i, value=val)
            total += val
        ws.cell(row=row_idx, column=33, value=total)
    
    for i in range(1, 34):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 8
    
    ws = wb.create_sheet(title='替代关系')
    ws.cell(row=1, column=1, value='')
    ws.merge_cells('B2:C2')
    ws.cell(row=2, column=2, value='原料替代关系')
    ws.merge_cells('K2:L2')
    ws.cell(row=2, column=11, value='替代数量上限')
    
    headers3 = ['', '序号', '替代类型', '', '替代物料一代码', '数量一', '替代物料二代码', '数量二', '比例', '整批'] + [str(i) for i in range(1, 29)]
    add_header(ws, 3, headers3)
    
    cursor = conn.execute('SELECT * FROM v_substitute')
    rows = cursor.fetchall()
    for row_idx, row_data in enumerate(rows, 4):
        ws.cell(row=row_idx, column=1, value=row_data['desc'] if row_data['desc'] else '')
        ws.cell(row=row_idx, column=2, value=row_data['seq'])
        ws.cell(row=row_idx, column=3, value=row_data['sub_type'])
        ws.cell(row=row_idx, column=4, value=row_data['material_type'])
        ws.cell(row=row_idx, column=5, value=row_data['code1'])
        ws.cell(row=row_idx, column=6, value=row_data['quantity1'])
        ws.cell(row=row_idx, column=7, value=row_data['code2'])
        ws.cell(row=row_idx, column=8, value=row_data['quantity2'])
        ws.cell(row=row_idx, column=9, value=row_data['ratio'])
        ws.cell(row=row_idx, column=10, value=row_data['batch'])
        for i in range(1, 29):
            col_name = f'limit_q{i}'
            val = row_data[col_name] if col_name in row_data.keys() else 0
            ws.cell(row=row_idx, column=10 + i, value=val)
    
    for i in range(1, 39):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 10
    
    wb.save(output_path)
    conn.close()
    print(f'✅ 数据已导出到: {output_path}')

if __name__ == '__main__':
    db_path = '/Users/duxiangyun/PythonProjects/APS/Excel2DB/data/aps_model.db'
    output_path = '/Users/duxiangyun/PythonProjects/APS/Excel2DB/aps_model_export.xlsx'
    export_to_excel(db_path, output_path)